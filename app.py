from flask import Flask, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import secrets
from sqlalchemy import func
import os
from werkzeug.utils import secure_filename
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from flask import send_file

app = Flask(__name__)

UPLOAD_FOLDER = 'static/uploads/comprobantes'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://u7r1sgp7zumpg5x4:1gFt1ukpHu89SQvvgiWV@bnqkyflugf81nksewztz-mysql.services.clever-cloud.com:3306/bnqkyflugf81nksewztz'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Clave secreta para sesiones (obligatoria para flash)
app.secret_key = secrets.token_hex(16)  # o usa una cadena fija como 'mi_clave_secreta_123'

db = SQLAlchemy(app)

# ---------- MODELOS ----------
class Producto(db.Model):
    __tablename__ = 'productos'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    precio_por_kilo = db.Column(db.Numeric(10,2), default=0.00)
    stock_actual = db.Column(db.Numeric(10,2), default=0.00)
    unidad = db.Column(db.String(20), default='kg')

class Compra(db.Model):
    __tablename__ = 'compras'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    proveedor = db.Column(db.String(150))
    total = db.Column(db.Numeric(10,2), default=0.00)
    observaciones = db.Column(db.Text)
    detalles = db.relationship('DetalleCompra', backref='compra', cascade='all, delete-orphan')

class DetalleCompra(db.Model):
    __tablename__ = 'detalle_compras'
    id = db.Column(db.Integer, primary_key=True)
    id_compra = db.Column(db.Integer, db.ForeignKey('compras.id'), nullable=False)
    id_producto = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=False)
    cantidad = db.Column(db.Numeric(10,2), nullable=False)
    precio_unitario = db.Column(db.Numeric(10,2), nullable=False)
    subtotal = db.Column(db.Numeric(10,2), nullable=False)
    producto = db.relationship('Producto')

class Venta(db.Model):
    __tablename__ = 'ventas'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    cliente = db.Column(db.String(150))
    telefono = db.Column(db.String(20))           # NUEVO
    total = db.Column(db.Numeric(10,2), default=0.00)
    observaciones = db.Column(db.Text)
    pagado = db.Column(db.Boolean, default=False) # NUEVO
    comprobante = db.Column(db.String(200))       # NUEVO (ruta del archivo)
    detalles = db.relationship('DetalleVenta', backref='venta', cascade='all, delete-orphan')

class DetalleVenta(db.Model):
    __tablename__ = 'detalle_ventas'
    id = db.Column(db.Integer, primary_key=True)
    id_venta = db.Column(db.Integer, db.ForeignKey('ventas.id'), nullable=False)
    id_producto = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=False)
    cantidad = db.Column(db.Numeric(10,2), nullable=False)
    precio_unitario = db.Column(db.Numeric(10,2), nullable=False)
    subtotal = db.Column(db.Numeric(10,2), nullable=False)
    producto = db.relationship('Producto')

# ---------- RUTAS ----------
@app.route('/')
def index():
    # Obtenemos algunos productos para mostrar en el panel (ejemplo)
    productos = Producto.query.limit(5).all()
    return render_template('index.html', productos=productos)

@app.route('/reportes')
def reportes():
    # Total de compras (convertir a float)
    total_compras = float(db.session.query(func.sum(Compra.total)).scalar() or 0.0)
    # Total de ventas (convertir a float)
    total_ventas = float(db.session.query(func.sum(Venta.total)).scalar() or 0.0)
    # Ganancia neta (ambos son float)
    ganancia_neta = total_ventas - total_compras

    # Número de compras y ventas
    num_compras = Compra.query.count()
    num_ventas = Venta.query.count()

    return render_template('reportes.html',
                           total_compras=total_compras,
                           total_ventas=total_ventas,
                           ganancia_neta=ganancia_neta,
                           num_compras=num_compras,
                           num_ventas=num_ventas)

@app.route('/reportes/compras')
def reporte_compras():
    # Obtener todas las compras con sus detalles
    compras = Compra.query.order_by(Compra.fecha.desc()).all()
    return render_template('reporte_compras.html', compras=compras)

@app.route('/reportes/ventas')
def reporte_ventas():
    ventas = Venta.query.order_by(Venta.fecha.desc()).all()
    return render_template('reporte_ventas.html', ventas=ventas)

@app.route('/ventas/<int:id>/pagar', methods=['GET', 'POST'])
def pagar_venta(id):
    venta = Venta.query.get_or_404(id)
    if request.method == 'POST':
        # Marcar como pagado
        venta.pagado = True
        # Subir comprobante
        if 'comprobante' in request.files:
            file = request.files['comprobante']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # Añadir timestamp para evitar duplicados
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                venta.comprobante = filename
        db.session.commit()
        flash('Pago registrado correctamente', 'success')
        return redirect(url_for('cronograma'))
    return render_template('pagar_venta.html', venta=venta)

@app.route('/cronograma')
def cronograma():
    ventas_pendientes = Venta.query.filter_by(pagado=False).order_by(Venta.fecha.asc()).all()
    return render_template('cronograma.html', ventas=ventas_pendientes)

@app.route('/productos')
def gestion_productos():
    productos = Producto.query.order_by(Producto.nombre).all()
    return render_template('productos.html', productos=productos)

@app.route('/productos/editar/<int:id>', methods=['GET', 'POST'])
def editar_producto(id):
    producto = Producto.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            producto.nombre = request.form.get('nombre', '').strip()
            producto.descripcion = request.form.get('descripcion', '').strip()
            producto.precio_por_kilo = float(request.form.get('precio', 0))
            # 🔽 CORRECCIÓN: tomar el stock del formulario, no usar 'cant'
            producto.stock_actual = float(request.form.get('stock', 0))
            db.session.commit()
            flash('Producto actualizado correctamente', 'success')
            return redirect(url_for('gestion_productos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}', 'danger')
            return render_template('editar_producto.html', producto=producto)
    
    return render_template('editar_producto.html', producto=producto)

@app.route('/productos/eliminar/<int:id>', methods=['POST'])
def eliminar_producto(id):
    try:
        producto = Producto.query.get_or_404(id)
        
        # 1. Eliminar detalles de compras asociados a este producto
        DetalleCompra.query.filter_by(id_producto=id).delete()
        
        # 2. Eliminar detalles de ventas asociados a este producto
        DetalleVenta.query.filter_by(id_producto=id).delete()
        
        # 3. Ahora eliminar el producto (ya no tiene dependencias)
        db.session.delete(producto)
        db.session.commit()
        
        flash('Producto eliminado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}', 'danger')
    return redirect(url_for('gestion_productos'))

@app.route('/configuracion', methods=['GET'])
def configuracion():
    return render_template('configuracion.html')

@app.route('/configuracion/exportar_excel')
def exportar_excel():
    # Crear un libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Encabezados
    headers = [
        "ID Venta", "Fecha", "Cliente", "Teléfono", 
        "Producto", "Cantidad (kg)", "Precio Unitario ($/kg)", 
        "Subtotal", "Total Venta", "Pagado", "Comprobante"
    ]
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Obtener todas las ventas con sus detalles
    ventas = Venta.query.order_by(Venta.fecha.desc()).all()
    
    row_num = 2
    for venta in ventas:
        # Si la venta tiene detalles, escribir una fila por cada detalle
        if venta.detalles:
            for detalle in venta.detalles:
                ws.cell(row=row_num, column=1, value=venta.id).border = border
                ws.cell(row=row_num, column=2, value=venta.fecha.strftime('%d/%m/%Y %H:%M')).border = border
                ws.cell(row=row_num, column=3, value=venta.cliente or '').border = border
                ws.cell(row=row_num, column=4, value=venta.telefono or '').border = border
                ws.cell(row=row_num, column=5, value=detalle.producto.nombre).border = border
                ws.cell(row=row_num, column=6, value=float(detalle.cantidad)).border = border
                ws.cell(row=row_num, column=7, value=float(detalle.precio_unitario)).border = border
                ws.cell(row=row_num, column=8, value=float(detalle.subtotal)).border = border
                ws.cell(row=row_num, column=9, value=float(venta.total)).border = border
                ws.cell(row=row_num, column=10, value="Sí" if venta.pagado else "No").border = border
                ws.cell(row=row_num, column=11, value=venta.comprobante or '').border = border
                row_num += 1
        else:
            # Si no tiene detalles (por si acaso), escribir una fila con datos básicos
            ws.cell(row=row_num, column=1, value=venta.id).border = border
            ws.cell(row=row_num, column=2, value=venta.fecha.strftime('%d/%m/%Y %H:%M')).border = border
            ws.cell(row=row_num, column=3, value=venta.cliente or '').border = border
            ws.cell(row=row_num, column=4, value=venta.telefono or '').border = border
            ws.cell(row=row_num, column=5, value="Sin productos").border = border
            ws.cell(row=row_num, column=6, value="").border = border
            ws.cell(row=row_num, column=7, value="").border = border
            ws.cell(row=row_num, column=8, value="").border = border
            ws.cell(row=row_num, column=9, value=float(venta.total)).border = border
            ws.cell(row=row_num, column=10, value="Sí" if venta.pagado else "No").border = border
            ws.cell(row=row_num, column=11, value=venta.comprobante or '').border = border
            row_num += 1

    # Ajustar anchos de columna automáticamente
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, row_num):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 4, 40)  # límite de 40 caracteres
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar en un buffer en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar el archivo como descarga
    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_ventas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/configuracion/vaciar-compras', methods=['POST'])
def vaciar_compras():
    try:
        # Eliminar todos los registros de compras y sus detalles (cascade lo hará)
        num_compras = Compra.query.count()
        Compra.query.delete()
        db.session.commit()
        flash(f'Historial de compras vaciado. Se eliminaron {num_compras} registros.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al vaciar compras: {str(e)}', 'danger')
    return redirect(url_for('index'))

@app.route('/configuracion/vaciar-ventas', methods=['POST'])
def vaciar_ventas():
    try:
        # Primero, eliminar los comprobantes asociados (archivos físicos)
        ventas = Venta.query.all()
        for venta in ventas:
            if venta.comprobante:
                try:
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], venta.comprobante))
                except:
                    pass  # Si no existe el archivo, continuar
        # Eliminar registros de ventas
        num_ventas = Venta.query.count()
        Venta.query.delete()
        db.session.commit()
        flash(f'Historial de ventas vaciado. Se eliminaron {num_ventas} registros y sus comprobantes.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al vaciar ventas: {str(e)}', 'danger')
    return redirect(url_for('index'))

@app.route('/configuracion/vaciar-comprobantes', methods=['POST'])
def vaciar_comprobantes():
    try:
        # Eliminar todos los archivos de comprobantes
        folder = app.config['UPLOAD_FOLDER']
        if os.path.exists(folder):
            archivos = os.listdir(folder)
            for archivo in archivos:
                try:
                    os.remove(os.path.join(folder, archivo))
                except:
                    pass
            flash(f'Comprobantes eliminados: {len(archivos)} archivos borrados.', 'success')
        else:
            flash('No hay comprobantes para eliminar.', 'info')
        
        # También limpiar la referencia en la BD (poner NULL)
        Venta.query.update({Venta.comprobante: None})
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error al vaciar comprobantes: {str(e)}', 'danger')
    return redirect(url_for('index'))

@app.route('/configuracion/vaciar-todo', methods=['POST'])
def vaciar_todo():
    try:
        # 1. Eliminar carpeta de comprobantes
        folder = app.config['UPLOAD_FOLDER']
        if os.path.exists(folder):
            shutil.rmtree(folder)
        os.makedirs(folder)

        # 2. Vaciar tablas con dependencias
        DetalleCompra.query.delete()
        Compra.query.delete()
        DetalleVenta.query.delete()
        Venta.query.delete()

        # 3. ELIMINAR TODOS LOS PRODUCTOS (STOCK BORRADO TOTALMENTE)
        Producto.query.delete()

        db.session.commit()
        flash('Todos los historiales, archivos, productos y stock han sido eliminados completamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al vaciar todo: {str(e)}', 'danger')
    return redirect(url_for('index'))

@app.route('/comprar', methods=['GET', 'POST'])
def comprar():
    if request.method == 'POST':
        # Obtener listas del formulario
        nombres = request.form.getlist('nombre[]')
        descripciones = request.form.getlist('descripcion[]')
        precios_compra = request.form.getlist('precio_compra[]')
        precios_venta = request.form.getlist('precio_venta[]')
        cantidades = request.form.getlist('cantidad[]')
        proveedor = request.form.get('proveedor', '').strip()
        observaciones = request.form.get('observaciones', '').strip()

        # Validar que al menos un producto tenga nombre
        if not nombres or not any(n.strip() for n in nombres):
            flash('Debe agregar al menos un producto', 'warning')
            return redirect(url_for('comprar'))

        total_compra = 0
        compra = Compra(proveedor=proveedor, observaciones=observaciones, total=0)
        db.session.add(compra)
        db.session.flush()  # para obtener el id

        for i in range(len(nombres)):
            nombre = nombres[i].strip()
            if not nombre:
                continue
            desc = descripciones[i].strip() if i < len(descripciones) else ''
            try:
                pc = float(precios_compra[i]) if precios_compra[i] else 0
            except:
                pc = 0
            try:
                pv = float(precios_venta[i]) if precios_venta[i] else 0
            except:
                pv = 0
            try:
                cant = float(cantidades[i]) if cantidades[i] else 0
            except:
                cant = 0

            if cant <= 0 or pc <= 0:
                continue  # omitir filas inválidas

            # Buscar o crear producto
            producto = Producto.query.filter_by(nombre=nombre).first()
            if not producto:
                producto = Producto(nombre=nombre, descripcion=desc, precio_por_kilo=pv, stock_actual=0)
                db.session.add(producto)
                db.session.flush()
            else:
                if desc:
                    producto.descripcion = desc
                if pv > 0:
                    producto.precio_por_kilo = pv

            # Crear detalle de compra
            subtotal = pc * cant
            detalle = DetalleCompra(
                id_compra=compra.id,
                id_producto=producto.id,
                cantidad=cant,
                precio_unitario=pc,
                subtotal=subtotal
            )
            db.session.add(detalle)

            # Actualizar stock
            producto.stock_actual = float(producto.stock_actual or 0) + cant
            total_compra += subtotal

        compra.total = total_compra
        db.session.commit()
        flash('Compra registrada exitosamente', 'success')
        return redirect(url_for('index'))

    # GET: mostrar formulario con una fila vacía
    return render_template('comprar.html')

@app.route('/vender', methods=['GET', 'POST'])
def vender():
    productos = Producto.query.filter(Producto.stock_actual > 0).order_by(Producto.nombre).all()
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        telefono = request.form.get('telefono', '').strip()
        cliente = f"{nombre} {apellido}".strip() if nombre or apellido else 'Cliente sin nombre'
        
        producto_ids = request.form.getlist('producto_id[]')
        cantidades = request.form.getlist('cantidad[]')
        observaciones = request.form.get('observaciones', '').strip()
        
        if not producto_ids or not any(pid.strip() for pid in producto_ids if pid):
            flash('Debe seleccionar al menos un producto y cantidad', 'warning')
            return redirect(url_for('vender'))
        
        total_venta = 0
        venta = Venta(cliente=cliente, telefono=telefono, total=0, observaciones=observaciones)
        db.session.add(venta)
        db.session.flush()
        
        for i in range(len(producto_ids)):
            pid = producto_ids[i].strip()
            if not pid:
                continue
            try:
                pid_int = int(pid)
            except:
                continue
            try:
                cant = float(cantidades[i]) if cantidades[i] else 0
            except:
                cant = 0
            if cant <= 0:
                continue
            producto = Producto.query.get(pid_int)
            if not producto:
                continue
            
            # 🔽 CORRECCIÓN: convertir stock a float antes de comparar y restar
            stock_float = float(producto.stock_actual)
            if stock_float < cant:
                flash(f'Stock insuficiente para {producto.nombre}. Disponible: {stock_float} kg', 'warning')
                db.session.rollback()
                return redirect(url_for('vender'))
            
            precio_unitario = float(producto.precio_por_kilo)
            subtotal = precio_unitario * cant
            detalle = DetalleVenta(
                id_venta=venta.id,
                id_producto=producto.id,
                cantidad=cant,
                precio_unitario=precio_unitario,
                subtotal=subtotal
            )
            db.session.add(detalle)
            # 🔽 AHORA AMBOS SON FLOAT
            producto.stock_actual = stock_float - cant
            total_venta += subtotal
        
        venta.total = total_venta
        db.session.commit()
        flash(f'Venta registrada exitosamente. Total: ${total_venta:.2f}', 'success')
        return redirect(url_for('index'))
    
    return render_template('vender.html', productos=productos)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)